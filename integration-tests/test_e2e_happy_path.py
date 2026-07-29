"""End-to-End 통합 테스트 — 실제 등록된 API 경로 기준

실행 전제:
- 7개 서비스가 8001~8007에서 기동 중 (start_all.ps1)
- tools/fixtures/master_sample.xlsx 존재

실행:
    pytest integration-tests/test_e2e_happy_path.py -v

경로 근거: tools/api_inventory.json (각 서비스 /openapi.json 실측)
계약 근거: bmad/00_SHARED_CONTRACT.md §5 — 모든 URL은 /api/v1 프리픽스,
           응답 봉투는 {"data": ..., "error": null}
"""
import time
from pathlib import Path

import httpx
import pytest

BASE = {
    "vm": "http://127.0.0.1:8001",
    "dist": "http://127.0.0.1:8002",
    "resp": "http://127.0.0.1:8003",
    "sched": "http://127.0.0.1:8004",
    "repair": "http://127.0.0.1:8005",
    "notify": "http://127.0.0.1:8006",
    "audit": "http://127.0.0.1:8007",
}

API = "/api/v1"
ROOT = Path(__file__).resolve().parent.parent
FIXTURE = ROOT / "tools" / "fixtures" / "master_sample.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def data_of(response):
    """계약 §5 봉투에서 data를 꺼낸다."""
    body = response.json()
    assert body.get("error") is None, f"error 필드가 비어있지 않음: {body.get('error')}"
    return body.get("data")


@pytest.fixture(scope="module", autouse=True)
def require_services():
    """7개 서비스가 모두 떠 있지 않으면 전체 모듈 skip."""
    down = []
    for name, base in BASE.items():
        try:
            if httpx.get(f"{base}/healthz", timeout=2.0).status_code != 200:
                down.append(name)
        except Exception:
            down.append(name)
    if down:
        pytest.skip(f"서비스 미기동: {', '.join(down)} — start_all.ps1 먼저 실행")


@pytest.fixture(scope="module")
def round_id():
    return f"R2026-E2E-{int(time.time()) % 100000}"


@pytest.fixture(scope="module")
def chain(round_id):
    """4개 호출이 넘겨주는 ID를 모듈 전체에서 공유."""
    return {"round_id": round_id}


# ---------------------------------------------------------------- Happy Path
# 4개 호출은 ID 체인이므로 순서 의존적이다. 앞 단계 실패 시 뒤는 skip 된다.

def test_01_register_master(chain):
    """[1/4] 마스터 등록 — multipart/form-data (JSON 아님)"""
    if not FIXTURE.exists():
        pytest.skip(f"픽스처 없음: {FIXTURE}")

    with FIXTURE.open("rb") as f:
        r = httpx.post(
            f"{BASE['vm']}{API}/versions/register",
            files={"file": (FIXTURE.name, f, XLSX_MIME)},
            data={"round_id": chain["round_id"], "kind": "master", "actor": "pytest-e2e"},
            timeout=30.0,
        )

    assert r.status_code in (200, 201), f"HTTP {r.status_code}: {r.text[:300]}"
    d = data_of(r)
    assert d["version_id"], "version_id 누락"
    assert d["applicant_count"] > 0, "지원자 0명 — 픽스처 파싱 실패"
    chain["version_id"] = d["version_id"]


def test_02_create_distribution_plan(chain):
    """[2/4] 배포안 생성 — master_version_id 필수"""
    if "version_id" not in chain:
        pytest.skip("1단계 실패 — version_id 없음")

    r = httpx.post(
        f"{BASE['dist']}{API}/distribute/plan",
        json={
            "round_id": chain["round_id"],
            "master_version_id": chain["version_id"],
            "created_by": "pytest-e2e",
        },
        timeout=60.0,
    )

    assert r.status_code in (200, 201), f"HTTP {r.status_code}: {r.text[:300]}"
    d = data_of(r)
    assert d["plan_id"], "plan_id 누락"
    assert d["total_applicants"] > 0
    assert d["team_counts"], "팀별 배정 결과 없음"
    chain["plan_id"] = d["plan_id"]


def test_03_approve_plan(chain):
    """[2.5/4] 배포안 승인 — 스케줄 생성의 선행 조건"""
    if "plan_id" not in chain:
        pytest.skip("2단계 실패 — plan_id 없음")

    r = httpx.post(
        f"{BASE['dist']}{API}/distribute/{chain['plan_id']}/approve",
        json={"actor": "pytest-e2e"},
        timeout=30.0,
    )

    assert r.status_code in (200, 201), f"HTTP {r.status_code}: {r.text[:300]}"
    d = data_of(r)
    assert d["status"] == "approved", f"승인 상태 아님: {d['status']}"


def test_04_generate_schedule(chain):
    """[3/4] 스케줄 생성 — plan_id 필수"""
    if "plan_id" not in chain:
        pytest.skip("2단계 실패 — plan_id 없음")

    r = httpx.post(
        f"{BASE['sched']}{API}/schedules/generate",
        json={
            "round_id": chain["round_id"],
            "plan_id": chain["plan_id"],
            "generated_by": "pytest-e2e",
        },
        timeout=120.0,
    )

    assert r.status_code in (200, 201), f"HTTP {r.status_code}: {r.text[:300]}"
    d = data_of(r)
    assert d["schedule_id"], "schedule_id 누락"
    assert d["total_assigned"] > 0, "배정 인원 0명"
    # 규칙 2는 HARD 제약 — 위반 시 스케줄 자체가 무효
    assert d["hard_violations"] == 0, f"HARD 규칙 위반 {d['hard_violations']}건"
    chain["schedule_id"] = d["schedule_id"]


def test_05_audit_timeline(chain):
    """[4/4] 감사 타임라인 조회 — 01/02/04 이벤트가 07에 도달했는지 확인

    포워딩(shared/contracts/audit_sink.py)은 백그라운드 스레드라 잠깐 재시도한다.
    """
    events = []
    for _ in range(10):
        r = httpx.get(
            f"{BASE['audit']}{API}/audit/timeline",
            params={"round_id": chain["round_id"]},
            timeout=15.0,
        )
        assert r.status_code == 200, f"HTTP {r.status_code}: {r.text[:300]}"
        events = data_of(r)
        assert isinstance(events, list), "data가 리스트가 아님"
        if events:
            break
        time.sleep(0.3)

    if "schedule_id" not in chain:
        pytest.skip("앞 단계 실패 — 검증할 이벤트 없음")
    types = {e.get("event_type") for e in events}
    assert "MASTER_REGISTERED" in types, f"01 이벤트 미도달: {types}"
    assert "DISTRIBUTION_PLAN_CREATED" in types, f"02 이벤트 미도달: {types}"
    assert "SCHEDULE_GENERATED" in types, f"04 이벤트 미도달: {types}"


def test_05b_schedule_traces_back_to_uploaded_file(chain):
    """시간표에 나온 사람이 1단계에 올린 그 파일 안에 실제로 있는지 확인.

    04가 목 데이터로 88명을 지어내면 여기서 잡힌다. 시간표의 이름을 원본
    엑셀에서 검색했을 때 안 나오는 상황을 막는 회귀 테스트다.
    """
    if "schedule_id" not in chain:
        pytest.skip("앞 단계 실패 — 대조할 시간표 없음")

    from openpyxl import load_workbook

    workbook = load_workbook(FIXTURE, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    header = {str(v).strip(): i for i, v in enumerate(rows[1]) if v}
    id_col = header["지원자 번호"]
    name_col = header["한글성명"]
    source = {
        str(r[id_col]).strip(): str(r[name_col]).strip()
        for r in rows[2:]
        if r[id_col] not in (None, "지원자 번호")
    }

    r = httpx.get(f"{BASE['sched']}{API}/schedules/{chain['schedule_id']}", timeout=30.0)
    assert r.status_code == 200, f"HTTP {r.status_code}: {r.text[:300]}"
    assignments = data_of(r)["assignments"]
    assert assignments, "배정 결과가 비어 있음"

    unknown = [a for a in assignments if str(a["applicant_id"]) not in source]
    assert not unknown, (
        f"원본 파일에 없는 지원자가 {len(unknown)}명 배정됨 "
        f"(예: {unknown[0]['applicant_id']} {unknown[0].get('applicant_name')})"
    )
    mismatched = [
        (a["applicant_id"], a.get("applicant_name"), source[str(a["applicant_id"])])
        for a in assignments
        if a.get("applicant_name") and source[str(a["applicant_id"])] != a["applicant_name"]
    ]
    assert not mismatched, f"지원자 번호는 있는데 이름이 다름: {mismatched[:3]}"


def test_06_dashboard_kpi(chain):
    """대시보드 KPI 조회 — 200 응답 확인"""
    r = httpx.get(
        f"{BASE['audit']}{API}/dashboard/kpi",
        params={"round_id": chain["round_id"]},
        timeout=15.0,
    )

    assert r.status_code == 200, f"HTTP {r.status_code}: {r.text[:300]}"
    assert data_of(r) is not None


# ------------------------------------------------------- 감사 이벤트 수집 경로
# fakeredis 는 프로세스 내부 전용이라 서비스별 uvicorn 프로세스 사이에서는
# Pub/Sub 이 닿지 않는다(mismatch_report.md M-8). 그래서 각 서비스가
# shared/contracts/audit_sink.py 로 07의 HTTP 수집 경로에 직접 넣는다.
# 아래는 그 수집 경로 자체의 왕복 검증.

def test_07_audit_direct_ingestion_roundtrip():
    """POST /audit/events 로 넣은 이벤트가 타임라인에 나타나는지 확인"""
    rid = f"R2026-INGEST-{int(time.time()) % 100000}"
    envelope = {
        "event_type": "MASTER_REGISTERED",
        "round_id": rid,
        "producer": "pytest-e2e",
        "payload": {"version_id": "vm_test", "applicant_count": 1},
    }

    # 비동기 수집이므로 202 Accepted가 정상 응답
    r = httpx.post(f"{BASE['audit']}{API}/audit/events", json=envelope, timeout=15.0)
    assert r.status_code in (200, 201, 202), f"HTTP {r.status_code}: {r.text[:300]}"
    assert data_of(r)["accepted"] == 1, "이벤트가 수집되지 않음"

    r = httpx.get(
        f"{BASE['audit']}{API}/audit/timeline", params={"round_id": rid}, timeout=15.0
    )
    assert r.status_code == 200
    events = data_of(r)
    assert len(events) >= 1, "직접 수집한 이벤트가 타임라인에 없음"
    assert events[0]["event_type"] == "MASTER_REGISTERED"


# ------------------------------------------------------------ 계약 준수 검증

@pytest.mark.parametrize("name,base", list(BASE.items()))
def test_08_contract_healthz(name, base):
    """계약 §5 — 모든 서비스가 GET /healthz 제공"""
    r = httpx.get(f"{base}/healthz", timeout=5.0)
    assert r.status_code == 200, f"{name}: HTTP {r.status_code}"


@pytest.mark.parametrize("name,base", list(BASE.items()))
def test_09_contract_api_v1_prefix(name, base):
    """계약 §5 — 업무 엔드포인트는 모두 /api/v1 프리픽스 사용"""
    spec = httpx.get(f"{base}/openapi.json", timeout=5.0).json()
    infra = {"/healthz", "/metrics", "/"}
    offenders = [
        p for p in spec.get("paths", {})
        if p not in infra and not p.startswith(API) and not p.startswith("/form/")
    ]
    assert not offenders, f"{name}: /api/v1 프리픽스 누락 → {offenders}"
