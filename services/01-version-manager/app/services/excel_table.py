"""엑셀 원본 표를 컬럼 손실 없이 읽고 다시 쓰는 유틸.

취합파일은 0행에 그룹 헤더(병합 셀), 1행에 실제 헤더가 오는 2단 구조이고
중간에 헤더 행이 그대로 반복되기도 한다. 병합 결과물은 생년월일·국적처럼
업무에서 안 쓰는 컬럼까지 원본 그대로 남겨야 하므로, Applicant 같은 도메인
모델을 거치지 않고 표 자체를 다룬다.

excel_parser 는 지원자 ID만 뽑는 가벼운 경로이고, 이 모듈은 셀 단위 대조와
병합 파일 생성을 위한 무거운 경로다. 헤더 탐색 규칙은 둘이 같다.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import Workbook, load_workbook

ID_HEADER = "지원자 번호"
NAME_HEADER = "한글성명"
HEADER_SCAN_ROWS = 15


class TableError(Exception):
    """표 구조를 알아볼 수 없을 때."""


def norm(value) -> str:
    """셀 값을 비교 가능한 문자열로 정규화한다.

    엑셀은 같은 숫자를 100001 / 100001.0 / '100001' 로 오갈 수 있어서,
    정규화 없이 비교하면 없는 차이가 잡힌다.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.lower() in ("nan", "none"):
        return ""
    if text.endswith(".0") and text[:-2].lstrip("-").isdigit():
        text = text[:-2]
    return text


@dataclass
class SheetTable:
    """한 시트의 헤더 + 데이터 행. 원본 셀 값을 그대로 들고 있는다."""

    sheet_name: str
    header_row: int
    preamble: list[list] = field(default_factory=list)  # 헤더 위쪽 행(그룹 헤더 등)
    header: list = field(default_factory=list)
    rows: list[list] = field(default_factory=list)
    id_col: int = 0

    @property
    def columns(self) -> dict[str, int]:
        """컬럼명 → 인덱스. 같은 이름이 뒤에 또 나오면 앞의 것을 쓴다.

        배포본에는 '지원자 번호'/'한글성명' 컬럼이 뒤쪽에 다시 붙어 있는데
        내용이 비어 있는 보조 컬럼이라 앞의 것이 정본이다.
        """
        out: dict[str, int] = {}
        for index, name in enumerate(self.header):
            key = norm(name)
            if key and key not in out:
                out[key] = index
        return out

    def row_map(self, row: list) -> dict[str, str]:
        return {name: norm(_at(row, index)) for name, index in self.columns.items()}

    def applicant_id(self, row: list) -> str:
        return norm(_at(row, self.id_col))

    def by_id(self) -> dict[str, list]:
        """지원자 번호 → 행. 같은 번호가 여러 번이면 첫 행을 쓴다."""
        out: dict[str, list] = {}
        for row in self.rows:
            aid = self.applicant_id(row)
            if aid and aid not in out:
                out[aid] = row
        return out

    def ids(self) -> list[str]:
        seen: list[str] = []
        known: set[str] = set()
        for row in self.rows:
            aid = self.applicant_id(row)
            if aid and aid not in known:
                known.add(aid)
                seen.append(aid)
        return seen


def _at(row: list, index: int):
    return row[index] if index < len(row) else None


def _locate_header(rows: list[list]) -> tuple[int, int]:
    for r in range(min(HEADER_SCAN_ROWS, len(rows))):
        for c, value in enumerate(rows[r]):
            if norm(value) == ID_HEADER:
                return r, c
    raise TableError(f"'{ID_HEADER}' 헤더를 찾지 못했습니다")


def read_table(source: bytes | str) -> SheetTable:
    """엑셀 바이트(또는 경로)에서 첫 시트를 표로 읽는다."""
    handle = BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
    workbook = load_workbook(handle, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        raw = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheet_name = sheet.title
    finally:
        workbook.close()

    header_row, id_col = _locate_header(raw)
    header = raw[header_row]
    rows = []
    for row in raw[header_row + 1:]:
        value = norm(_at(row, id_col))
        if not value or value == ID_HEADER:
            # 반복 헤더 행 · 빈 행 · 소계 행은 데이터가 아니다
            continue
        rows.append(row)
    return SheetTable(
        sheet_name=sheet_name,
        header_row=header_row,
        preamble=raw[:header_row],
        header=header,
        rows=rows,
        id_col=id_col,
    )


def write_table(table: SheetTable, rows: list[list]) -> bytes:
    """원본 레이아웃(그룹 헤더 + 헤더 + 데이터)을 유지한 채 새 파일을 만든다."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = table.sheet_name or "master"
    for line in table.preamble:
        sheet.append(list(line))
    sheet.append(list(table.header))
    width = len(table.header)
    for row in rows:
        padded = list(row)[:width]
        padded += [None] * (width - len(padded))
        sheet.append(padded)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()
