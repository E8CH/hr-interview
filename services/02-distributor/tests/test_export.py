"""팀별 엑셀 내보내기 테스트."""
import io

import pytest
from openpyxl import load_workbook

BODY = {"round_id": "R2026-Q3-01", "master_version_id": "vm_abc123"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def plan_id(client):
    return client.post("/api/v1/distribute/plan", json=BODY).json()["data"]["plan_id"]


def test_export_team_excel(client, plan_id):
    response = client.get(f"/api/v1/distribute/{plan_id}/export/AI솔루션팀")
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MIME

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    assert sheet.title == "AI솔루션팀"
    header = [cell.value for cell in sheet[1]]
    assert header[:2] == ["지원자 번호", "한글성명"]
    assert "배포사유태그" in header
    # 정원 16명 + 중복 배포분
    assert sheet.max_row >= 17

    tag_col = header.index("배포사유태그") + 1
    for row in range(2, sheet.max_row + 1):
        tags = sheet.cell(row=row, column=tag_col).value
        assert tags and len(tags.split(", ")) >= 2


def test_export_all_teams(client, plan_id):
    for team in ["AI솔루션팀", "로봇응용기술팀", "미래혁신팀", "배터리기술팀", "전극기술팀"]:
        response = client.get(f"/api/v1/distribute/{plan_id}/export/{team}")
        assert response.status_code == 200, team


def test_export_unknown_team(client, plan_id):
    response = client.get(f"/api/v1/distribute/{plan_id}/export/없는팀")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "NOT_FOUND"


def test_export_unknown_plan(client):
    response = client.get("/api/v1/distribute/nope/export/AI솔루션팀")
    assert response.status_code == 404


def test_export_file_persisted_to_storage(client, plan_id):
    from app.config import settings

    client.get(f"/api/v1/distribute/{plan_id}/export/전극기술팀")
    path = settings.storage_dir / plan_id / "전극기술팀.xlsx"
    assert path.exists() and path.stat().st_size > 0
