"""팀별 엑셀 내보내기 (openpyxl) — PoC는 로컬 ./storage/ 저장."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings

HEADERS = [
    ("applicant_id", "지원자 번호", 14),
    ("name", "한글성명", 12),
    ("team_1st", "1지망_조직", 14),
    ("job_1st", "1지망_직무", 14),
    ("degree", "학위구분", 10),
    ("major_final", "최종학력_주전공", 20),
    ("major_bachelor", "학사1_주전공", 20),
    ("gpa_final", "환산학점", 10),
    ("target_lab", "타겟랩여부", 12),
    ("advisor", "지도교수", 12),
    ("score", "배포점수", 10),
    ("tags", "배포사유태그", 42),
    ("duplicate", "중복배포", 10),
    ("primary_team", "주관팀", 16),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_DUP_FILL = PatternFill("solid", fgColor="FFF2CC")


def _degree_label(degree_type: str | None) -> str:
    return "학사" if (degree_type or "").strip() == "과정1" else "대학원"


def build_team_workbook(plan_id: str, team_name: str, rows: list[dict]) -> Workbook:
    """assignment_reasons 행 목록 → 워크북."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = team_name[:31]

    sheet.append([label for _key, label, _width in HEADERS])
    for idx, (_key, _label, width) in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(idx)].width = width

    for row in rows:
        snapshot = row.get("snapshot") or {}
        sheet.append(
            [
                row.get("applicant_id"),
                row.get("applicant_name") or snapshot.get("name"),
                snapshot.get("team_1st"),
                snapshot.get("job_1st"),
                _degree_label(snapshot.get("degree_type")),
                snapshot.get("major_final"),
                snapshot.get("major_bachelor"),
                snapshot.get("gpa_final"),
                snapshot.get("target_lab"),
                snapshot.get("advisor"),
                row.get("score"),
                ", ".join(row.get("tags") or []),
                "Y" if row.get("is_duplicate") else "N",
                row.get("primary_team") or "",
            ]
        )
        if row.get("is_duplicate"):
            for idx in range(1, len(HEADERS) + 1):
                sheet.cell(row=sheet.max_row, column=idx).fill = _DUP_FILL

    sheet.freeze_panes = "A2"
    return workbook


def export_team_excel(plan_id: str, team_name: str, rows: list[dict]) -> Path:
    """워크북을 STORAGE_DIR에 저장하고 경로를 반환."""
    workbook = build_team_workbook(plan_id, team_name, rows)
    out_dir: Path = settings.storage_dir / plan_id
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{team_name}.xlsx"
    workbook.save(path)
    return path
