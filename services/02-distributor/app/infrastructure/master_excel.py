"""마스터 취합파일(.xlsx) 파서.

실제 취합파일 구조:
- 1행: 그룹 헤더('희망조직 및 직무', '학교, 학력' …) → 무시
- 2행: 실제 컬럼 헤더
- 3행부터 데이터 (중간에 헤더가 반복되는 행이 섞여 있어 걸러낸다)
- 뒤쪽에 '지원자 번호'·'한글성명' 컬럼이 중복 등장 → 첫 번째 것만 사용
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from contracts.types import Applicant
from openpyxl import load_workbook

HEADER_ROW = 1  # 0-based (엑셀 2행)

#: Applicant 필드 → 엑셀 컬럼명
COLUMN_MAP = {
    "applicant_id": "지원자 번호",
    "name": "한글성명",
    "team_1st": "1지망_조직",
    "job_1st": "1지망_직무",
    "rnd_type": "R&D/N-R&D",
    "degree_type": "최종학력_학교유형",
    "major_final": "최종학력_주전공",
    "major_bachelor": "학사1_주전공",
    "gpa_final": "최종학력_환산학점",
    "target_lab": "타겟랩여부",
    "advisor": "지도교수",
    "prev_applications": "이전 지원 전체 횟수",
    "doc_result": "1차서류 결과",
}

REQUIRED_COLUMNS = ("지원자 번호", "1지망_조직", "1지망_직무", "R&D/N-R&D", "1차서류 결과")

#: 01(병합)이 팀별 배포본에서 읽어 새겨 둔 팀 나눔. 한 사람을 두 팀이 적어 냈으면
#: 쉼표로 이어 붙어 있다. 필수는 아니다 — 이 컬럼이 없던 시절의 취합파일도 읽는다.
TEAM_COLUMN = "담당팀"


class MasterFileError(Exception):
    """취합파일 구조가 예상과 다름."""


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _split_teams(value) -> list[str]:
    """'AI솔루션팀, 전극기술팀' → ['AI솔루션팀', '전극기술팀'] (순서·중복 정리)."""
    text = _clean(value)
    if not text:
        return []
    out: list[str] = []
    for piece in text.split(","):
        team = piece.strip()
        if team and team not in out:
            out.append(team)
    return out


def _to_int(value) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def load_master_excel(source: str | Path | bytes) -> list[Applicant]:
    """취합파일을 Applicant 리스트로 파싱한다.

    경로뿐 아니라 바이트도 받는다 — 01(version-manager)에서 원본 파일을
    내려받아 그대로 파싱하는 경로가 있어서다.
    """
    if isinstance(source, (bytes, bytearray)):
        handle = BytesIO(source)
        path = "<uploaded bytes>"
    else:
        path = Path(source)
        if not path.exists():
            raise MasterFileError(f"취합파일이 없습니다: {path}")
        handle = path

    workbook = load_workbook(handle, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if len(rows) <= HEADER_ROW:
        raise MasterFileError(f"데이터가 없습니다: {path}")

    header = rows[HEADER_ROW]
    index: dict[str, int] = {}
    for position, label in enumerate(header):
        label = _clean(label)
        if label and label not in index:  # 중복 컬럼은 첫 번째만
            index[label] = position

    missing = [c for c in REQUIRED_COLUMNS if c not in index]
    if missing:
        raise MasterFileError(f"필수 컬럼 누락: {missing}")

    id_col = index["지원자 번호"]
    applicants: list[Applicant] = []
    seen: set[str] = set()

    for row in rows[HEADER_ROW + 1 :]:
        applicant_id = _clean(row[id_col]) if id_col < len(row) else None
        if not applicant_id or applicant_id == "지원자 번호":
            continue  # 빈 행 · 반복 헤더 행
        if applicant_id in seen:
            continue
        seen.add(applicant_id)

        def cell(field: str):
            position = index.get(COLUMN_MAP[field])
            if position is None or position >= len(row):
                return None
            return row[position]

        team_col = index.get(TEAM_COLUMN)
        teams = (
            _split_teams(row[team_col])
            if team_col is not None and team_col < len(row)
            else []
        )

        applicants.append(
            Applicant(
                applicant_id=applicant_id,
                name=_clean(cell("name")) or "",
                team_1st=_clean(cell("team_1st")) or "",
                job_1st=_clean(cell("job_1st")) or "",
                rnd_type=_clean(cell("rnd_type")) or "",
                degree_type=_clean(cell("degree_type")) or "",
                major_final=_clean(cell("major_final")),
                major_bachelor=_clean(cell("major_bachelor")),
                gpa_final=_to_float(cell("gpa_final")),
                target_lab=_clean(cell("target_lab")),
                advisor=_clean(cell("advisor")),
                prev_applications=_to_int(cell("prev_applications")),
                doc_result=_clean(cell("doc_result")) or "",
                assigned_teams=teams,
            )
        )

    if not applicants:
        raise MasterFileError(f"파싱된 지원자가 0명입니다: {path}")
    return applicants


def load_team_roster(path: str | Path) -> list[str]:
    """희망지원자_{팀}.xlsx → 배포된 지원자 번호 목록 (백테스트 정답지)."""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    ids: list[str] = []
    for row in rows[HEADER_ROW + 1 :]:
        applicant_id = _clean(row[0]) if row else None
        if not applicant_id or applicant_id == "지원자 번호":
            continue
        ids.append(applicant_id)
    return ids
