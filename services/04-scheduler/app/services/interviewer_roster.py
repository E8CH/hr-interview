"""면접관명단.xlsx 업로드 · 회차별 선별.

콘솔 3번 메뉴(면접 담당자 선별)가 쓰는 경로다.
  - 업로드: 사번 | 성명 | 직급 | 소속팀 | 이메일 | 일일최대 | 우선순위 | 가능시간
            을 마스터에 반영
  - 선별: 회차마다 그중 누구를 투입할지 골라 round_interviewers 에 저장

시간 단위 가용성(어느 요일 몇 시)은 03(회신 수집)이 모은다. 다만 현업이 명단을
낼 때 "앞타임 / 뒤타임" 정도는 미리 적어 두므로, 가능시간 컬럼이 있으면 그
덩어리를 요일 전체에 펼쳐 마스터 가용성으로 저장한다 — 회신이 오면 04가
둘을 교집합으로 합친다.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.domain.interviewer import Interviewer
from app.domain.round_interviewer import RoundInterviewer
from app.errors import ValidationFailed
from app.infrastructure.contracts import (
    BAND_ALL,
    BAND_BACK,
    BAND_FRONT,
    BAND_NONE,
    HOURS,
    TIME_BANDS,
    band_availability,
    band_hours,
    band_name,
    band_of,
)

# 엑셀 컬럼명 → 모델 필드. 운영에서 쓰는 표기 흔들림을 흡수한다.
COLUMN_ALIASES = {
    "interviewer_id": ("사번", "사원번호", "면접관ID", "interviewer_id"),
    "name": ("성명", "이름", "한글성명", "name"),
    "title": ("직급", "직위", "직책", "title"),
    "team": ("소속팀", "팀", "소속", "team"),
    "email": ("이메일", "메일", "email"),
    "max_daily": ("일일최대", "하루최대", "max_daily"),
    "priority": ("우선순위", "priority"),
    "time_band": ("가능시간", "가능 시간", "오전오후", "앞뒤타임", "time_band"),
}
REQUIRED = ("interviewer_id", "team")
HEADER_SCAN_ROWS = 10
DEFAULT_MAX_DAILY = len(HOURS)
DEFAULT_PRIORITY = 2


def _norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return "" if text.lower() in ("nan", "none") else text


def _to_int(value, default: int) -> int:
    text = _norm(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def _band(value) -> str:
    """적어 낸 가능시간을 앞타임 · 뒤타임 표기로 정리한다 (빈칸이면 빈 문자열).

    현업이 내는 엑셀에는 '오전만' 처럼 예전 표기가 그대로 적혀 오는 일이 많다.
    오전 = 앞타임, 오후 = 뒤타임으로 받아 준다.
    """
    text = _norm(value).replace(" ", "")
    if not text:
        return ""
    named = band_name(text)
    if named in TIME_BANDS:
        return named
    has_front = "오전" in text or "앞" in text
    has_back = "오후" in text or "뒤" in text
    if has_front and has_back:
        return BAND_ALL
    if has_front:
        return BAND_FRONT
    if has_back:
        return BAND_BACK
    if text in ("종일", "하루종일", "전일", "모두", "전체", "둘다"):
        return BAND_ALL
    if text in ("불가", "없음", "어렵다", "-"):
        return BAND_NONE
    return ""


def _locate_header(rows: list[list]) -> tuple[int, dict[str, int]]:
    """어떤 행이 헤더인지 별칭 표로 찾아낸다."""
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        labels = {_norm(v): c for c, v in enumerate(row) if _norm(v)}
        mapping = {}
        for field, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in labels:
                    mapping[field] = labels[alias]
                    break
        if all(field in mapping for field in REQUIRED):
            return index, mapping
    raise ValidationFailed(
        "면접관명단 헤더를 찾지 못했습니다 — 사번 · 소속팀 컬럼이 필요합니다"
    )


def parse_roster(data: bytes) -> list[dict]:
    """엑셀 바이트 → 면접관 dict 목록 (사번 기준 중복 제거)."""
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationFailed(f"엑셀을 열 수 없습니다: {exc}") from exc
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    header_row, columns = _locate_header(rows)

    def cell(row: list, field: str):
        index = columns.get(field)
        return row[index] if index is not None and index < len(row) else None

    out: list[dict] = []
    seen: set[str] = set()
    for row in rows[header_row + 1:]:
        interviewer_id = _norm(cell(row, "interviewer_id"))
        team = _norm(cell(row, "team"))
        if not interviewer_id or interviewer_id in seen:
            continue
        if interviewer_id in [a for aliases in COLUMN_ALIASES.values() for a in aliases]:
            continue  # 반복된 헤더 행
        if not team:
            continue
        seen.add(interviewer_id)
        out.append({
            "interviewer_id": interviewer_id,
            "name": _norm(cell(row, "name")),
            "title": _norm(cell(row, "title")),
            "team": team,
            "email": _norm(cell(row, "email")),
            "max_daily": _to_int(cell(row, "max_daily"), DEFAULT_MAX_DAILY),
            "priority": _to_int(cell(row, "priority"), DEFAULT_PRIORITY),
            "time_band": _band(cell(row, "time_band")),
        })
    if not out:
        raise ValidationFailed("면접관 행을 하나도 읽지 못했습니다")
    return out


def import_roster(db: Session, data: bytes) -> dict:
    """업로드한 명단을 마스터에 반영한다 (사번 기준 upsert).

    가능시간 칸을 적어 냈을 때만 가용성을 새로 쓴다 — 빈칸이면 이미 회신으로
    채워진 값을 그대로 둔다.
    """
    parsed = parse_roster(data)
    created = updated = 0
    for row in parsed:
        fields = {k: v for k, v in row.items() if k != "time_band"}
        band = row.get("time_band") or ""
        existing = db.get(Interviewer, row["interviewer_id"])
        if existing is None:
            db.add(Interviewer(
                **fields, availability=band_availability(band) if band else {}
            ))
            created += 1
        else:
            for field, value in fields.items():
                if field != "interviewer_id":
                    setattr(existing, field, value)
            if band:
                existing.availability = band_availability(band)
            updated += 1
    db.commit()
    teams = sorted({row["team"] for row in parsed})
    return {
        "parsed": len(parsed),
        "created": created,
        "updated": updated,
        "teams": teams,
        "interviewers": parsed,
    }


def set_bands(db: Session, bands: dict[str, str], actor: str = "console",
              timing: dict | None = None) -> dict:
    """사번별 가능 시간(앞타임 · 뒤타임)을 마스터에 반영한다.

    고른 덩어리가 곧 하루에 볼 수 있는 최대 인원의 천장이 된다 — 뒤타임만 되는
    사람에게 하루 여덟 명을 배정할 수는 없으므로 일일최대를 칸 수까지 줄인다.

    어느 칸이 앞타임이고 어느 칸이 뒤타임인지는 면접 진행 조건이 정한다.
    두 덩어리는 12시 ~ 14시에서 겹치므로 그 사이의 칸은 양쪽 다 맡을 수 있다.
    예전 표기('오전만' 등)로 들어와도 지금 이름으로 옮겨 받는다.
    """
    bands = {iid: band_name(band) for iid, band in (bands or {}).items()}
    bad = [b for b in bands.values() if b not in TIME_BANDS]
    if bad:
        raise ValidationFailed(
            f"알 수 없는 가능 시간입니다: {bad[0]}", {"allowed": list(TIME_BANDS)}
        )
    rows = {
        r.interviewer_id: r
        for r in db.scalars(
            select(Interviewer).where(Interviewer.interviewer_id.in_(list(bands)))
        ).all()
    } if bands else {}
    unknown = [i for i in bands if i not in rows]
    if unknown:
        raise ValidationFailed(
            f"등록되지 않은 면접관입니다: {', '.join(unknown[:5])}",
            {"unknown": unknown},
        )

    changed = 0
    for interviewer_id, band in bands.items():
        row = rows[interviewer_id]
        hours = band_hours(band, timing)
        if row.availability and sorted(
            {h for day_hours in row.availability.values() for h in day_hours}
        ) == sorted(hours):
            continue  # 바뀐 게 없으면 손대지 않는다 (직접 줄여 둔 일일최대 보존)
        row.availability = band_availability(band, timing=timing)
        row.max_daily = len(hours)
        changed += 1
    if changed:
        db.commit()
    return {"updated": changed, "actor": actor}


def select_for_round(db: Session, round_id: str, interviewer_ids: list[str],
                     actor: str = "console") -> dict:
    """회차 선별 목록을 통째로 교체한다."""
    known = set(db.scalars(
        select(Interviewer.interviewer_id).where(
            Interviewer.interviewer_id.in_(interviewer_ids)
        )
    ).all()) if interviewer_ids else set()
    unknown = [i for i in interviewer_ids if i not in known]
    if unknown:
        raise ValidationFailed(
            f"등록되지 않은 면접관입니다: {', '.join(unknown[:5])}",
            {"unknown": unknown},
        )

    db.execute(delete(RoundInterviewer).where(RoundInterviewer.round_id == round_id))
    for interviewer_id in dict.fromkeys(interviewer_ids):
        db.add(RoundInterviewer(
            round_id=round_id, interviewer_id=interviewer_id, selected_by=actor
        ))
    db.commit()
    return {"round_id": round_id, "selected": len(set(interviewer_ids))}


def selected_ids(db: Session, round_id: str) -> list[str]:
    return list(db.scalars(
        select(RoundInterviewer.interviewer_id)
        .where(RoundInterviewer.round_id == round_id)
        .order_by(RoundInterviewer.interviewer_id)
    ).all())


def list_for_round(db: Session, round_id: str) -> list[dict]:
    ids = selected_ids(db, round_id)
    if not ids:
        return []
    rows = db.scalars(
        select(Interviewer).where(Interviewer.interviewer_id.in_(ids))
        .order_by(Interviewer.team, Interviewer.priority, Interviewer.interviewer_id)
    ).all()
    return [
        {
            "interviewer_id": r.interviewer_id,
            "name": r.name,
            "title": r.title or "",
            "team": r.team,
            "email": r.email,
            "max_daily": r.max_daily,
            "priority": r.priority,
            "availability": r.availability or {},
            # 화면은 시간 한 칸씩이 아니라 앞타임 · 뒤타임 덩어리로 읽는다.
            # 아직 아무것도 적지 않았으면 빈칸 — 배치할 때는 하루 종일로 본다.
            "time_band": band_of(r.availability) if r.availability else "",
        }
        for r in rows
    ]
